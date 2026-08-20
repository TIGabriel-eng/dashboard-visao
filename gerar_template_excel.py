import os
import sys
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'visao_academy.settings')
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
django.setup()

from openpyxl import Workbook
from core.models import Perfil

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = 'Usuários'

# Cabeçalho
headers = ['username', 'email', 'first_name', 'last_name', 'role', 
           'empresa', 'cnpj', 'telefone', 'cargo', 'unidade', 
           'regime_federal', 'is_empresario']
ws.append(headers)

# Exemplos de dados
exemplos = [
    ['joao.silva', 'joao@empresa.com', 'João', 'Silva', 'cliente_vex', 
     'Empresa ABC', '12.345.678/0001-90', '(11)98765-4321', 'Analista', 'sao_paulo', 'mei', True],
    
    ['maria.souza', 'maria@empresa.com', 'Maria', 'Souza', 'empresário', 
     'Empresa XYZ', '98.765.432/0001-10', '(11)91234-5678', 'Diretora', 'salvador', 'me', True],
    
    ['pedro.costa', 'pedro@empresa.com', 'Pedro', 'Costa', 'cliente_equipe', 
     'Empresa 123', '11.222.333/0001-44', '(11)99999-8888', 'Gerente', 'maracas', '', False],
    
    ['ana.oliveira', 'ana@empresa.com', 'Ana', 'Oliveira', 'colaborador_vex', 
     'Visão Tributária', '00.000.000/0001-00', '(71)3333-4444', 'Colaborador', '', '', False],
]

for exemplo in exemplos:
    ws.append(exemplo)

# Adicionar segunda planilha com roles válidas
ws_roles = wb.create_sheet('Roles Válidas')
ws_roles.append(['Role', 'Descrição'])
for role, descricao in Perfil.ROLE_CHOICES:
    ws_roles.append([role, descricao])

# Adicionar terceira planilha com unidades válidas
ws_unidades = wb.create_sheet('Unidades Válidas')
ws_unidades.append(['Unidade', 'Descrição'])
for unidade, descricao in Perfil.UNIDADE_CHOICES:
    if unidade:  # Ignorar opção vazia
        ws_unidades.append([unidade, descricao])

# Adicionar quarta planilha com regimes federais
ws_regimes = wb.create_sheet('Regimes Federais')
ws_regimes.append(['Regime', 'Descrição'])
for regime, descricao in Perfil.REGIME_FEDERAL_CHOICES:
    if regime:  # Ignorar opção vazia
        ws_regimes.append([regime, descricao])

# Salvar
output_path = 'template_usuarios_massa.xlsx'
wb.save(output_path)
print(f'Template criado com sucesso: {output_path}')
print(f'Localização: {os.path.abspath(output_path)}')