from io import BytesIO
from datetime import datetime

from django.contrib.auth.models import User

from core.models import Perfil

COLUNAS = [
    'username', 'email', 'first_name', 'last_name', 'role',
    'empresa', 'cnpj', 'telefone', 'cargo', 'unidade',
    'regime_federal', 'is_empresario',
]

COLUNAS_OBRIGATORIAS = ['username', 'email']


def gerar_senha(first_name, last_name):
    base = f"{first_name}.{last_name}".lower()
    base = ''.join(c for c in base if c.isalnum() or c == '.')
    return f"{base}@123"


def validar_role(role):
    roles_validas = [r[0] for r in Perfil.ROLE_CHOICES]
    if role not in roles_validas:
        return None
    return role


def processar_arquivo_excel(arquivo, planilha_nome='Sheet1'):
    import openpyxl

    wb = openpyxl.load_workbook(arquivo)
    ws = wb[planilha_nome] if planilha_nome in wb.sheetnames else wb.active

    headers = [cell.value for cell in ws[1]]

    col_map = {}
    for col in COLUNAS:
        if col in headers:
            col_map[col] = headers.index(col)

    erros_map = {}
    for obrigatoria in COLUNAS_OBRIGATORIAS:
        if obrigatoria not in col_map:
            erros_map[obrigatoria] = f'Coluna obrigatória "{obrigatoria}" não encontrada no cabeçalho.'

    if erros_map:
        return None, erros_map

    usuarios_criados = []
    erros = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        linha_vazia = all(cell is None for cell in row)
        if linha_vazia:
            continue

        try:
            username = row[col_map['username']]
            email = row[col_map['email']]
            first_name = row[col_map.get('first_name')] or '' if 'first_name' in col_map else ''
            last_name = row[col_map.get('last_name')] or '' if 'last_name' in col_map else ''
            role = row[col_map.get('role')] or 'cliente_orcoma' if 'role' in col_map else 'cliente_orcoma'
            empresa = row[col_map.get('empresa')] or '' if 'empresa' in col_map else ''
            cnpj = row[col_map.get('cnpj')] or '' if 'cnpj' in col_map else ''
            telefone = row[col_map.get('telefone')] or '' if 'telefone' in col_map else ''
            cargo = row[col_map.get('cargo')] or '' if 'cargo' in col_map else ''
            unidade = row[col_map.get('unidade')] or '' if 'unidade' in col_map else ''
            regime_federal = row[col_map.get('regime_federal')] or '' if 'regime_federal' in col_map else ''
            is_empresario_raw = row[col_map.get('is_empresario')] if 'is_empresario' in col_map else False
            if isinstance(is_empresario_raw, str):
                is_empresario = is_empresario_raw.lower() in ('true', '1', 'sim', 'yes')
            else:
                is_empresario = bool(is_empresario_raw)

            if not username or not email:
                erros.append({'linha': idx, 'username': str(username or ''), 'motivo': 'username ou email vazio.'})
                continue

            if User.objects.filter(username=username).exists():
                erros.append({'linha': idx, 'username': username, 'motivo': 'Usuário já existe.'})
                continue

            role_valida = validar_role(role)
            if not role_valida:
                erros.append({'linha': idx, 'username': username, 'motivo': f'Role inválida: "{role}".'})
                continue

            senha = gerar_senha(first_name or username, last_name or 'user')

            user = User.objects.create_user(
                username=username,
                email=email,
                password=senha,
                first_name=first_name,
                last_name=last_name,
            )

            perfil = user.perfil
            perfil.role = role_valida
            perfil.empresa = empresa
            perfil.cnpj = cnpj
            perfil.telefone = telefone
            perfil.cargo = cargo
            perfil.unidade = unidade
            perfil.regime_federal = regime_federal
            perfil.is_empresario = is_empresario
            perfil.save()

            usuarios_criados.append({
                'username': username,
                'email': email,
                'password': senha,
                'role': role_valida,
                'empresa': empresa,
                'status': 'Criado',
            })

        except Exception as e:
            erros.append({'linha': idx, 'username': str(row[col_map.get('username', 0)] or '?') if col_map.get('username') is not None else '?', 'motivo': str(e)})

    timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    return {
        'usuarios': usuarios_criados,
        'erros': erros,
        'total_processado': len(usuarios_criados) + len(erros),
        'sucessos': len(usuarios_criados),
        'total_erros': len(erros),
        'timestamp': timestamp,
    }, None


def gerar_template_bytes():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Usuários'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1a3d6e', end_color='1a3d6e', fill_type='solid')

    for col_idx, header in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for col_idx in range(1, len(COLUNAS) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 20

    ws_roles = wb.create_sheet('Roles Válidas')
    ws_roles.append(['Role', 'Descrição'])
    for role, descricao in Perfil.ROLE_CHOICES:
        ws_roles.append([role, descricao])

    ws_unidades = wb.create_sheet('Unidades Válidas')
    ws_unidades.append(['Unidade', 'Descrição'])
    for unidade, descricao in Perfil.UNIDADE_CHOICES:
        if unidade:
            ws_unidades.append([unidade, descricao])

    ws_regimes = wb.create_sheet('Regimes Federais')
    ws_regimes.append(['Regime', 'Descrição'])
    for regime, descricao in Perfil.REGIME_FEDERAL_CHOICES:
        if regime:
            ws_regimes.append([regime, descricao])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def gerar_relatorio_bytes(usuarios_criados):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Usuários Cadastrados'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1a3d6e', end_color='1a3d6e', fill_type='solid')

    cabecalhos = ['Username', 'Email', 'Senha', 'Role', 'Empresa', 'Status']
    for col_idx, header in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for col_idx in range(1, len(cabecalhos) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 22

    for u in usuarios_criados:
        ws.append([
            u['username'],
            u['email'],
            u['password'],
            u['role'],
            u['empresa'],
            u['status'],
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
