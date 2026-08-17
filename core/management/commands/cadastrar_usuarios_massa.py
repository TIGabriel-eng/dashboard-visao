import os
import sys
import django
from datetime import datetime

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'visao_academy.settings')
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
django.setup()

from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from core.models import Perfil
import openpyxl
from openpyxl import Workbook


class Command(BaseCommand):
    help = 'Cadastra usuários em massa a partir de um arquivo Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            '--arquivo',
            type=str,
            required=True,
            help='Caminho para o arquivo Excel com os dados dos usuários'
        )
        parser.add_argument(
            '--planilha',
            type=str,
            default='Sheet1',
            help='Nome da planilha (padrão: Sheet1)'
        )

    def gerar_senha(self, first_name, last_name):
        """Gera senha no formato: Nome.Sobrenome@123"""
        base = f"{first_name}.{last_name}".lower()
        base = ''.join(c for c in base if c.isalnum() or c == '.')
        senha = f"{base}@123"
        return senha

    def validar_role(self, role):
        """Valida se a role é válida"""
        roles_validas = [r[0] for r in Perfil.ROLE_CHOICES]
        if role not in roles_validas:
            return None
        return role

    def handle(self, *args, **options):
        arquivo_path = options['arquivo']
        planilha_nome = options['planilha']

        if not os.path.exists(arquivo_path):
            self.stdout.write(self.style.ERROR(f'Arquivo não encontrado: {arquivo_path}'))
            return

        try:
            wb = openpyxl.load_workbook(arquivo_path)
            ws = wb[planilha_nome] if planilha_nome in wb.sheetnames else wb.active
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Erro ao abrir arquivo: {str(e)}'))
            return

        # Ler cabeçalho
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        self.stdout.write(f'Cabeçalhos encontrados: {headers}')

        # Mapear colunas
        colunas_esperadas = ['username', 'email', 'first_name', 'last_name', 'role', 
                            'empresa', 'cnpj', 'telefone', 'cargo', 'unidade', 
                            'regime_federal', 'is_empresario']
        
        col_map = {}
        for col in colunas_esperadas:
            if col in headers:
                col_map[col] = headers.index(col)
        
        if 'username' not in col_map or 'email' not in col_map:
            self.stdout.write(self.style.ERROR('Colunas obrigatórias não encontradas: username, email'))
            return

        # Processar linhas
        total_linhas = ws.max_row - 1
        sucessos = 0
        erros = 0
        usuarios_criados = []

        self.stdout.write(f'Processando {total_linhas} usuários...')

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Extrair dados
                username = row[col_map['username']]
                email = row[col_map['email']]
                first_name = row[col_map.get('first_name', 0)] or ''
                last_name = row[col_map.get('last_name', 0)] or ''
                role = row[col_map.get('role', 0)] or 'cliente_orcoma'
                empresa = row[col_map.get('empresa', 0)] or ''
                cnpj = row[col_map.get('cnpj', 0)] or ''
                telefone = row[col_map.get('telefone', 0)] or ''
                cargo = row[col_map.get('cargo', 0)] or ''
                unidade = row[col_map.get('unidade', 0)] or ''
                regime_federal = row[col_map.get('regime_federal', 0)] or ''
                is_empresario = row[col_map.get('is_empresario', 0)] or False

                # Validações básicas
                if not username or not email:
                    self.stdout.write(self.style.WARNING(f'Linha {idx}: username ou email vazio. Ignorando.'))
                    erros += 1
                    continue

                # Verificar se usuário já existe
                if User.objects.filter(username=username).exists():
                    self.stdout.write(self.style.WARNING(f'Linha {idx}: Usuário {username} já existe. Ignorando.'))
                    erros += 1
                    continue

                # Validar role
                role_valida = self.validar_role(role)
                if not role_valida:
                    self.stdout.write(self.style.WARNING(f'Linha {idx}: Role inválida "{role}". Ignorando.'))
                    erros += 1
                    continue

                # Gerar senha
                senha = self.gerar_senha(first_name or username, last_name or 'user')

                # Criar usuário
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=senha,
                    first_name=first_name,
                    last_name=last_name
                )

                # Atualizar perfil
                perfil = user.perfil
                perfil.role = role_valida
                perfil.empresa = empresa
                perfil.cnpj = cnpj
                perfil.telefone = telefone
                perfil.cargo = cargo
                perfil.unidade = unidade
                perfil.regime_federal = regime_federal
                perfil.is_empresario = is_empresario
                perfil.save()

                # O signal atribuir_plano_por_cnpj vai atribuir o plano automaticamente

                sucessos += 1
                usuarios_criados.append({
                    'username': username,
                    'email': email,
                    'password': senha,
                    'role': role_valida,
                    'empresa': empresa,
                    'status': 'Criado'
                })

                if idx % 100 == 0:
                    self.stdout.write(f'  Processados {idx}/{total_linhas}...')

            except Exception as e:
                self.stdout.write(self.style.ERROR(f'Linha {idx}: Erro ao processar: {str(e)}'))
                erros += 1

        # Gerar relatório Excel
        if usuarios_criados:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            relatorio_path = f'relatorio_usuarios_cadastrados_{timestamp}.xlsx'
            
            wb_relatorio = Workbook()
            ws_relatorio = wb_relatorio.active
            ws_relatorio.title = 'Usuários Cadastrados'
            
            # Cabeçalho
            ws_relatorio.append(['Username', 'Email', 'Senha', 'Role', 'Empresa', 'Status'])
            
            # Dados
            for u in usuarios_criados:
                ws_relatorio.append([
                    u['username'],
                    u['email'],
                    u['password'],
                    u['role'],
                    u['empresa'],
                    u['status']
                ])
            
            wb_relatorio.save(relatorio_path)
            self.stdout.write(self.style.SUCCESS(f'\nRelatório salvo em: {relatorio_path}'))

        # Resumo
        self.stdout.write(self.style.SUCCESS(f'\n=== RESUMO ==='))
        self.stdout.write(f'Total processado: {total_linhas}')
        self.stdout.write(self.style.SUCCESS(f'Sucessos: {sucessos}'))
        self.stdout.write(self.style.WARNING(f'Erros: {erros}'))
        self.stdout.write(f'\nArquivo de entrada: {arquivo_path}')
        if usuarios_criados:
            self.stdout.write(f'Arquivo de saída: relatorio_usuarios_cadastrados_{timestamp}.xlsx')